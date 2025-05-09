import time
import praw
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font


# Function to create a Reddit instance
def create_reddit_instance(client_id, client_secret, user_agent):
    reddit = praw.Reddit(
        client_id=client_id,
        client_secret=client_secret,
        user_agent=user_agent
    )
    return reddit


def load_submission_data_to_table(subreddit, keywords, tags, amount, sort_type, time_filter, time_limit):
    rows = []
    post_number = 1
    added_rows = 0
    match sort_type:
        case 'hot':
            ranking = subreddit.hot(limit=None)
        case 'new':
            ranking = subreddit.new(limit=None)
        case 'rising':
            ranking = subreddit.rising(limit=None)
        case 'top':
            ranking = subreddit.top(time_filter=time_filter, limit=None)
        case _:
            print("invalid sort_type")
            return
    try:
        start_time = time.time()
        prev_minutes, prev_seconds = -1, -1
        for submission in ranking:
            elapsed_time = time.time() - start_time
            minutes, seconds = divmod(int(elapsed_time), 60)
            if prev_minutes != minutes or prev_seconds != seconds:
                print(f"Time passed: {minutes:02}:{seconds:02}")
            prev_minutes, prev_seconds = minutes, seconds
            if added_rows >= amount:
                print("Amount reached... stopping search.")
                return rows
            if elapsed_time > time_limit != -1:
                print("Time limit reached... stopping search.")
                return rows
            if submission.author == 'PokeUpdateBot':
                continue
            if submission.link_flair_text not in tags and len(tags) > 0:
                continue
            rows.append({
                "Username": str(submission.author),
                "Content": f'Post {post_number}|Tag: {submission.link_flair_text}|-- {submission.title}:\n{submission.selftext.replace('\n', ' ')}',
                "URL": submission.url,
                "Depth": 0
            })
            post_number += 1
            submission.comments.replace_more(limit=0)  # Load all comments
            # Process top-level comments
            comment_number = 1  # Reset comment number for each post
            for top_level_comment in submission.comments:
                current_rows, has_keyword = (
                    extract_comments(submission, top_level_comment, keywords,
                                     depth=2, comment_number=comment_number))
                if has_keyword:
                    comment_number += 1
                    rows.extend(current_rows)
            if comment_number == 1:
                rows.pop(-1)
            else:
                rows.append({
                    "Username": "",
                    "Content": "",
                    "Depth": 0
                })
                rows.append({
                    "Username": "",
                    "Content": "",
                    "Depth": 0
                })
                added_rows += 1
            print(f'Posts added: {added_rows}')
        print("All posts searched... stopping search.")
    finally:
        return rows


def extract_comments(submission, comment, keywords, depth=1, comment_number=1):
    comments_data = []
    indent_type = "Comment" if depth > 1 else "Post"
    indent = f'{indent_type} {comment_number} | '
    indented_comment = '    ' * (depth - 1) + indent + comment.body.replace('\n', ' ')  # Remove unnecessary line breaks
    comments_data.append({
        "Username": str(comment.author),
        "Content": indented_comment,
        "URL": f"https://www.reddit.com{submission.permalink}{comment.id}",
        "Depth": depth
    })
    if len(comment.replies) == 0:
        if len(keywords) == 0:
            return comments_data, True
        for keyword in keywords:
            if keyword.lower() in comment.body.lower():
                return comments_data, True
    reply_number = 1
    has_keyword = False
    for reply in comment.replies:
        current_row, has_keyword = extract_comments(submission, reply, keywords, depth + 1, reply_number) # Recurse into replies
        if has_keyword:
            comments_data.extend(current_row)
        reply_number += 1
    return comments_data, has_keyword


# Function to search Reddit posts and save them with comments in a tree structure to an Excel file
def save_data_to_xlsx(subreddit, keywords, tags, amount, sort_type, time_filter, time_limit, filename="reddit_posts.xlsx"):
    rows = load_submission_data_to_table(subreddit, keywords, tags, amount, sort_type, time_filter, time_limit)
    df = pd.DataFrame(rows)
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Content"])
    colors = ['F0F0F0', 'D3D3D3', 'B0C4DE', 'ADD8E6', 'E6E6FA', 'FFFACD', 'F5DEB3', 'FAFAD2', 'E0FFFF', 'F5F5F5']
    for _, row in df.iterrows():
        depth = row["Depth"]
        color_index = depth % len(colors)
        if depth == 0:
            ws.append([row["Username"], row["Content"], row["URL"]])
        else:
            fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
            ws.append([row["Username"], row["Content"], row["URL"]])
            ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)  # Ensure text is wrapped nicely

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    make_keyword_cells_bold_in_cells(keywords, ws)
    wb.save(filename)
    print(f"Data saved to {filename}")


def make_keyword_cells_bold_in_cells(keywords, ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                # Check each keyword
                for keyword in keywords:
                    if keyword.lower() in cell.value.lower() and keyword != "":
                        # Change the entire cell to bold
                        cell.font = Font(bold=True)
                        break  # Exit after the first keyword is found


def get_client_credentials():
    with open('client_credentials.txt', 'r') as file:
        for line in file:
            name, value = line.strip().split('=')
            if name == 'client_id':
                client_id = value
            elif name == 'client_secret':
                client_secret = value
        return client_id, client_secret


def main():
    # Reddit API credentials
    client_id, client_secret = get_client_credentials()
    user_agent = "reddit search by crusader"
    reddit = create_reddit_instance(client_id, client_secret, user_agent)

    subreddit = reddit.subreddit('pokemon')
    keywords = ['totodile']  # '[keyword]', add all the keywords you want to track, leave empty for no keyword search
    tags = []  # Options: '[tag]', add all the tags you want to track, leave empty for no tag search
    sort_type = 'top'  # Options: 'hot', 'new', 'rising', 'top'
    time_filter = 'week'  # Options (only works with sort_type = 'top'): "all", "year", "month", "week", "day", "hour"
    amount = 10  # Amount of posts that are being saved
    time_limit = -1  # Time limit for search in seconds, value -1 is infinite

    save_data_to_xlsx(
        subreddit=subreddit,
        keywords=keywords,
        tags=tags,
        amount=amount,
        sort_type=sort_type,
        time_filter=time_filter,
        time_limit=time_limit,
        filename="reddit_posts.xlsx")


if __name__ == "__main__":
    main()

